# Run: Python3 sar.py
# From New_Santander -> sar.py

from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl.worksheet.dimensions
from enum import Enum
from utils import Utils

IP_FILENAME = 'presub.xlsx'
IP_SHEETNAME = 'Sheet1'
OP_FILENAME = 'BMW.xlsx'
OP_SHEETNAME = 'Sheet1'

wb1 = load_workbook(IP_FILENAME)
wsA = wb1[IP_SHEETNAME]

wb2 = Workbook()
wsB = wb2.active
wsB.title = OP_SHEETNAME
wsB = wb2[OP_SHEETNAME]

MAX_COLUMN = 20
MIN_ROW = 1
MAX_ROW = 200 # 65

# Offset for merged headers
OFFSET = 0

i = 1
# Copy over First Row
for col in Utils.bmw_columns:
    wsB.cell(row=1, column=i).value = col
    i += 1

MIN_ROW += 1

# clean up the output sheet
# for row in wsB.iter_rows(min_row=MIN_ROW + OFFSET, max_col=45, max_row=MAX_ROW):
#     for cell in row:
#         cell.value = None

# Options for data copy
Copy = Enum('Copy', [('NONE', 0), ('VALUE', 1), ('CIA', 2), ('DATE', 3)])
# None if set value = None
# Value if copy value
# Yes if value = 'Y'
# No if value = 'N'
# DSAR if value = 'DSAR'
# CIA if Third Party Name

# {Copy.CIA -> a}
# {A -> b, E -> c, F -> e}
# {G -> f, H -> g, I -> h}
# {J -> o, K -> p, L -> q, M -> r}


# Company Name
# Record ID, Agreement Number, Vehicle Registration
# Fname, Lname, DOB
# Address, Postcode, Prev Address, Prev Postcode
# LOA URL
mappings = [['B', 'a', Copy.CIA], 
            ['A', 'b', Copy.VALUE], ['D', 'c', Copy.VALUE], ['E', 'e', Copy.VALUE], 
            ['F', 'f', Copy.VALUE], ['G', 'g', Copy.VALUE], ['H', 'h', Copy.DATE], 
            ['I', 'o', Copy.VALUE], ['J', 'p', Copy.VALUE], ['K', 'q', Copy.VALUE], ['L', 'r', Copy.VALUE],
            ['M', 's', Copy.VALUE]]

seen = set()

# range -> [a, b)
for row in range(MIN_ROW, MAX_ROW):

    # Guards
    if wsA.cell(row=row, column=1).value is None: 
        break

    fan = wsA.cell(row=row, column=Utils.toIndex('D')).value
    if fan in seen:
        # Print Dupes
        print(fan)
        continue

    seen.add(fan)

    for mapping in mappings:
        colA = Utils.toIndex(mapping[0])
        colB = Utils.toIndex(mapping[1])
        match mapping[2]:
            case Copy.NONE:
                wsB.cell(row=row + OFFSET, column=colB).value = None
            case Copy.VALUE:
                wsB.cell(row=row + OFFSET, column=colB).value = wsA.cell(row=row, column=colA).value
            case Copy.CIA:
                wsB.cell(row=row + OFFSET, column=colB).value = 'Consultation Claims Ltd'
            case Copy.DATE:
                wsB.cell(row=row + OFFSET, column=colB).value = wsA.cell(row=row, column=colA).value
                wsB.cell(row=row + OFFSET, column=colB).number_format = 'DD/MM/YYYY'
            case _:
                continue

# for col in wsB.columns:
#     wsB.column_dimensions[col[0].column_letter].auto_size = True

for letter in range(1, 10):
    if wsB.cell(row=2, column=letter).value != None :
        wsB.column_dimensions[Utils.toLetter(letter)].width = 20

for letter in range(15, 20):
    if wsB.cell(row=2, column=letter).value != None :
        if letter % 2 == 1:
            wsB.column_dimensions[Utils.toLetter(letter)].width = 50
        else :
            wsB.column_dimensions[Utils.toLetter(letter)].width = 15

# Save and Close
wb2.save(OP_FILENAME)
wb1.close()
wb2.close()