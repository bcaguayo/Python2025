# Run: Python3 nodupes.py
# >>> exec(open('nodupes.py').read())

import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

FILENAME = 'Reengagements.xlsx'
SHEETNAME = 'HubSpot-2'

wb1 = load_workbook(FILENAME)
wsA = wb1[SHEETNAME]

NUM_ROWS = 1000
NUM_COLS = 13
DEAL_COUNT_COL = 3

# Count the number of non-empty cells in column 14 if not duplicated
seen = dict()
for row in range(2, NUM_ROWS):  # 2 to 105
    fname = wsA.cell(row=row, column=1).value
    lname = wsA.cell(row=row, column=2).value
    if fname == None or lname == None: continue
    name = fname + ' ' + lname
    if name not in seen:
        seen.update({name:row})
        print(f"Adding {name} to dict at row {row}")
    elif name in seen:
        # Increase Count of Deals
        srcRow = seen[name]
        count = wsA.cell(row=srcRow, column=DEAL_COUNT_COL).value
        wsA.cell(row=srcRow, column=DEAL_COUNT_COL).value = count + 1
        print(f"Duplicate {name} found at row {row}, increasing count to {count + 1}")
        # Erase Row
        for col in range(1, NUM_COLS + 1):
            # Delete the row
            # print(f"Deleting {name} at row {row}")
            # wsA.delete_rows(row, 1)
            wsA.cell(row=row, column=col).value = None

print("Number of unique non-empty cells:", len(seen))

wb1.save(FILENAME)
wb1.close()