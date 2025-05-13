# Run: Python3 nodupes.py
# >>> exec(open('nodupes.py').read())

import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

FILENAME = 'Reengagements.xlsx'
SHEETNAME = 'HubSpot'

wb1 = load_workbook(FILENAME)
wsA = wb1[SHEETNAME]

NUM_ROWS = 1000
NUM_COLS = 13

# Count the number of non-empty cells in column 14 if not duplicated
row = 2
inarrow = 0
while row <= NUM_ROWS and inarrow < 100:
    if wsA.cell(row=row, column=1).value is None:
        wsA.delete_rows(row, 1)
        inarrow += 1
    else:
        row += 1
        inarrow = 0

print(f"Deleted {row} empty rows, {inarrow} foam")

wb1.save(FILENAME)
wb1.close()