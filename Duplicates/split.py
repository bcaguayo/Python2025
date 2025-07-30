# Run: Python3 parts.py

# Import math library
import math
import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

# init values
NUM_COLUMNS = 15
NUM_ROWS = 3130

# Input -> filename and sheet name
FILENAME = 'split.xlsx'
SHEETNAME = 'src'
PART1 = 'live'
PART2 = 'lost'

# indexing
DEALNAME_COL = 10
offset1 = 2
offset2 = 2

# Open Workbook
wb = load_workbook(FILENAME)
ws = wb[SHEETNAME]
wl1 = wb[PART1]
wl2 = wb[PART2]

seen = ""

# Split data into n partitions
for row in range (2, NUM_ROWS + 1):
    name = ws.cell(row=row, column=DEALNAME_COL).value
    name = name.split('-')[0].strip()
    if name != seen:
        # Copy to Live
        for column in range (1, NUM_COLUMNS):
            wl1.cell(row=offset1, column=column).value = ws.cell(row=row, column=column).value
        offset1 += 1
        seen = name
    else:
        # Copy to Lost
        for column in range (1, NUM_COLUMNS):
            wl2.cell(row=offset2, column=column).value = ws.cell(row=row, column=column).value
        # Next Line
        offset2 += 1

# Womp womp
wb.save(FILENAME)
wb.close