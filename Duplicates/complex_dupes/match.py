# Duplicated Finance agreement numbers, with Tolerance
# Using sequential matching, no hash needed.

from openpyxl import load_workbook

# 1. From same contact
# 2. Keep cache of FAN
# 3. Split string by space or dash
# 4. If next contact, forget
# 5. Repeat

FILENAME = 'all.xlsx'
wb = load_workbook(FILENAME)
ws_active = wb.active

DEAL_COL = 1
CONTACT_COL = 2
FAN_COL = 3

rows = 2
contact = str(ws_active.cell(row=rows, column=CONTACT_COL).value).strip(' ')

# next cell not empty
while contact != '' and contact != 'None':
    fan = str(ws_active.cell(row=rows, column=FAN_COL).value).strip(' ')
    split_fan = fan.split(" ")[0]

    # Get params
    rows += 1
    contact = str(ws_active.cell(row=rows, column=CONTACT_COL).value).strip(' ')
    