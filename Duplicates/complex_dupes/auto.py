from openpyxl import load_workbook
import csv
import re

# === CONFIG ===
FILENAME = 'Duplicate Deals.xlsx'              # Excel file name (must be in same folder as this script)
CONTACT_ID_HEADER = 'Contact ID'               # Header name for Contact ID
FAN_HEADER = 'Finance Agreement Number'        # Header name for Finance Agreement
OUTPUT_FILE = 'duplicates_flagged_sorted.csv'  # Output CSV file name

# === Helper: normalize FANs ===
def normalize_fan(fan):
    """Treats '12345' and '12345 1' as the same base agreement."""
    if not fan or str(fan).strip().lower() == 'none':
        return ''
    fan = str(fan).strip()
    match = re.match(r'^(\d+)', fan)
    return match.group(1) if match else fan

# === Load workbook ===
wb = load_workbook(FILENAME)
ws = wb.active

# === Step 1: Locate column positions by header names ===
headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

try:
    CONTACT_ID_COL = headers.index(CONTACT_ID_HEADER) + 1
    FAN_COL = headers.index(FAN_HEADER) + 1
except ValueError as e:
    missing = CONTACT_ID_HEADER if CONTACT_ID_HEADER not in headers else FAN_HEADER
    raise ValueError(f"❌ Column header '{missing}' not found in the sheet. Check header names.") from e

print(f"✅ Found columns — Contact ID: {CONTACT_ID_COL}, FAN: {FAN_COL}")

# === Step 2: Identify duplicates ===
seen = {}
duplicates = set()

for i in range(2, ws.max_row + 1):
    contact_id = str(ws.cell(row=i, column=CONTACT_ID_COL).value or '').strip()
    fan_raw = str(ws.cell(row=i, column=FAN_COL).value or '').strip()
    fan_clean = normalize_fan(fan_raw)

    if not contact_id or not fan_clean:
        continue

    key = (contact_id, fan_clean)
    if key in seen:
        duplicates.add(key)
    else:
        seen[key] = i

# === Step 3: Collect all rows and flag duplicates ===
rows_data = []
header = headers + ['Duplicate_Flag']

for i in range(2, ws.max_row + 1):
    row_values = [cell.value for cell in ws[i]]
    contact_id = str(ws.cell(row=i, column=CONTACT_ID_COL).value or '').strip()
    fan_raw = str(ws.cell(row=i, column=FAN_COL).value or '').strip()
    fan_clean = normalize_fan(fan_raw)

    flag = 'Duplicate' if (contact_id, fan_clean) in duplicates else ''
    rows_data.append(row_values + [flag])

# === Step 4: Sort so duplicates appear first ===
rows_data.sort(key=lambda x: (x[-1] != 'Duplicate',))

# === Step 5: Write to CSV ===
with open(OUTPUT_FILE, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(header)
    writer.writerows(rows_data)

print(f"✅ Done! Sorted CSV with duplicate flags created: {OUTPUT_FILE}")
wb.close()
