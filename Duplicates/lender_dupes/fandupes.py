# Duplicated Finance agreement numbers

# Use hash to store seen fan and output duplicates

from openpyxl import load_workbook

FILENAME = 'all.xlsx'

wb = load_workbook(FILENAME)

ws_active = wb.active

# 22 tabs

seen = set()
num_dupes = 0
duplicates = set()

i = 1
cell = str(ws_active.cell(row=i, column=3).value).strip(' ')

# next cell not empty
while cell != '' and cell != 'None':
    if cell in seen and cell != '(No value)':
        num_dupes += 1
        record = str(ws_active.cell(row=i, column=4).value).strip(' ')
        duplicates.add(cell)
    else:
        seen.add(cell)

    # move to next cell
    i += 1
    cell = str(ws_active.cell(row=i, column=3).value).strip(' ')

print(f'Number of duplicates: {num_dupes}')

# Hash table
hash_table = {}

for i in range(1, ws_active.max_row + 1):
    cell = str(ws_active.cell(row=i, column=3).value).strip(' ')
    if cell in duplicates:
        record = str(ws_active.cell(row=i, column=4).value).strip(' ')
        hash_table.setdefault(cell, []).append(record)

# User input
print('Choose mode 1 for inline print, choose 2 for all deals print')
input_mode = input('Enter mode: ')
print(f'selected: {input_mode}')

def print_inline_duplicates(hash_table, filename='dupes_1.txt'):
    with open(filename, 'w') as f:
        for fan in hash_table:
            f.write(f'{fan}, {hash_table[fan]}\n')

def print_deals_only(hash_table, filename='dupes_2.txt'):
    with open(filename, 'w') as f:
        for fan in hash_table:
            for record in hash_table[fan]:
                f.write(record + '\n')

if input_mode == '1':
    print_inline_duplicates(hash_table)
elif input_mode == '2':
    print_deals_only(hash_table)
elif input_mode == '3':
    print_inline_duplicates(hash_table, 'dupes_1.txt')
    print_deals_only(hash_table, 'dupes_2.txt')
else:
    print('Wrong input, please try again')

wb.close