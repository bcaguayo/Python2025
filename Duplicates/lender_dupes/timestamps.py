# Duplicated Finance agreement numbers
# Divide list A into B and C
# Deterministic

from openpyxl import load_workbook
from datetime import datetime
from typing import Dict, List, Tuple

# 1 is FAN (not unique, match)
# 2 is Record Id (unique)
# 3 is Timestamp (compare)

FILENAME = 'timestamps.xlsx'

wb = load_workbook(FILENAME)
ws_active = wb.active

# Live container of oldest datapoint
data: Dict[str, Tuple[str, datetime]] = {}
# All younger 
dupes = set()

# Loop through filled rows
for row in ws_active.iter_rows(min_row=2, values_only=True):
    # Get params
    fan = str(row[0]).strip() if row[0] is not None else ''
    rec = str(row[1]).strip() if row[1] is not None else ''
    cell_value = row[2] if len(row) > 2 else None
    if isinstance(cell_value, datetime):
        timestamp = cell_value
    elif isinstance(cell_value, str):
        try:
            timestamp = datetime.fromisoformat(cell_value)
        except ValueError:
            timestamp = datetime.strptime(cell_value, "%Y-%m-%d %H:%M:%S")
    else:
        timestamp = None

    # Guards
    if not fan or not rec or timestamp is None:
        continue

    if fan == '' or fan == 'None':
        break

    # If agreement already in data
    if fan in data:
        existing_rec, existing_timestamp = data[fan]
        # Keep older agreement, put newer in dupes (rec only)
        if timestamp < existing_timestamp:
            dupes.add(existing_rec)
            data[fan] = (rec, timestamp)            
            # print(f'found duplicate at {fan}, kept {rec} over {existing_rec}')
        else:
            dupes.add(rec)            
            # print(f'found duplicate at {fan}, sent {rec} to dupes')    
    # Otherwise, add to data
    else:
        data[fan] = (rec, timestamp)

print('______________________________________\n')

# Extract all recs from data into a set
deals = {rec for rec, _ in data.values()}

wb.close

# Write B (originals) into filekeep, C (duplicates) into filedump
with open('filekeep.txt', 'w') as f:
    for deal in deals:
        f.write(deal + '\n')

with open('filedump.txt', 'w') as f:
    for deal in dupes:
        f.write(deal + '\n')