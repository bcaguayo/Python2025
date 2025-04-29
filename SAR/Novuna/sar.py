# Run: Python3 sar.py

import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

wb1 = load_workbook('SAR_Novuna.xlsx')
wsA = wb1['SAR_Novuna']

wb2 = load_workbook('Novuna.xlsx')
wsB = wb2['Sheet1']

# clean up the output sheet
for row in wsB.iter_rows(min_row=2, max_col=15, max_row=180):
    for cell in row:
        cell.value = None

# for i in range(1, 11):
#     wsB[f"A{i}"] = wsA[f"A{i}"].value

# [A1, B2, C3, D4, E5, F6, G7, H8, I9, J10, K11]
# [L12, M13, N14, O15, P16, Q17, R18, S19, T20, U21, V22, W23, X24]

# range -> [a, b)
for row in range(2, 176):
    # REF NUMBER
    for column in range(1, 2):  # A
        wsB.cell(row=row, column=column).value = wsA.cell(row=row, column=column).value
    # NAME
    for column in range(2, 4):  # A to B
        wsB.cell(row=row, column=column).value = wsA.cell(row=row, column=column + 2).value
    # PREV NAMES
    for column in range(4, 5):  # C to E
        wsB.cell(row=row, column=column).value = None
    # DOB
    for column in range(5, 6):  # F
        wsB.cell(row=row, column=column).value = wsA.cell(row=row, column=column + 1).value
    # ADDRESS
    for column in range(7, 9):
        # i = 0 for first column, i = 1 for second column
        i = column - 7
        fulladdress = wsA.cell(row=row, column=column).value
        # split on the last two commas
        split = fulladdress.rsplit(',', 2)

        if fulladdress is not None and len(split) >= 3:
            # Address
            wsB.cell(row=row, column=column - 1 + 2 * i).value = split[0]
            # Postcode
            wsB.cell(row=row, column=column + 2 * i).value = split[1]
            wsB.cell(row=row, column=column + 1 + 2 * i).value = split[2]
        elif column == 8 and len(split) < 3:
            wsB.cell(row=row, column=column + 1).value = wsB.cell(row=row, column=column - 2).value
            wsB.cell(row=row, column=column + 2).value = wsB.cell(row=row, column=column - 1).value
            wsB.cell(row=row, column=column + 3).value = wsB.cell(row=row, column=column).value
    # FAN 
    for column in range(9, 11):
        cell = wsA.cell(row=row, column=column).value
        if cell is not None:
            wsB.cell(row=row, column = column + 3).value = cell
        # otherwise, leave it blank
    for column in range(11, 12):
        wsB.cell(row=row, column = column + 3).value = "Y"

# wb3 = Workbook() # type: ignore
# wsC = wb3.active
# wsC.title = 'output'
wb2.save('Novuna.xlsx')