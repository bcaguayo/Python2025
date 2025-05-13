# Run: Python3 sar.py

from openpyxl import load_workbook # type: ignore
from enum import Enum
from utils import Utils

IP_FILENAME = 'SAR_Close_Brothers.xlsx'
IP_SHEETNAME = 'SAR_Close_Brothers'
OP_FILENAME = 'CloseBrothers.xlsx'
OP_SHEETNAME = 'Bulk Submission'

wb1 = load_workbook(IP_FILENAME)
wsA = wb1[IP_SHEETNAME]

wb2 = load_workbook(OP_FILENAME)
wsB = wb2[OP_SHEETNAME]

MIN_ROW = 2
MAX_ROW = 80

# Offset for merged headers
OFFSET = 0

# clean up the output sheet
for row in wsB.iter_rows(min_row=MIN_ROW + OFFSET, max_col=45, max_row=MAX_ROW):
    for cell in row:
        cell.value = None

# Options for data copy
Copy = Enum('Copy', [('NONE', 0), ('VALUE', 1), ('YES', 2), ('NO', 3), ('DSAR', 4)])
# None if set value = None
# Value if copy value
# Yes if value = 'Y'
# No if value = 'N'
# DSAR if value = 'DSAR'

# [Column on IP Spreadsheet, Column on OP Spreadsheet, Copy Enum]

# Novuna
# mappings = [['A', 'a', Copy.VALUE], ['D', 'b', Copy.VALUE], ['E', 'c', Copy.VALUE], ['N', 'd', Copy.NONE],
#            ['F', 'e', Copy.VALUE], ['I', 'l', Copy.VALUE], ['J', 'm', Copy.VALUE], ['N', 'n', Copy.YES]]
# streets = [['G', ['f', 'g', 'h']], ['H', ['i', 'j', 'k']]]

# Motonovo
# mappings = [['A', 'a', Copy.VALUE], ['A', 'b', Copy.DSAR], ['D', 'd', Copy.VALUE], ['E', 'e', Copy.VALUE],
#            ['F', 'f', Copy.VALUE], ['G', 'h', Copy.VALUE], ['I', 'm', Copy.VALUE], ['J', 'n', Copy.VALUE]]
# streets = [['H', ['i', 'j', 'l']], ['K', ['r', 's', 'u']]]

# CB
mappings = [['D', 'a', Copy.VALUE], ['E', 'b', Copy.VALUE], ['F', 'i', Copy.VALUE], ['J', 'an', Copy.VALUE], 
            ['K', 'ao', Copy.VALUE], ['L', 'ap', Copy.VALUE], ['A', 'aq', Copy.VALUE]]
streets = [['G', ['j', 'l', 'n']], ['H', ['o', 'q', 's']]]

seen = set()

# range -> [a, b)
for row in range(MIN_ROW, MAX_ROW):

    # Guards
    if wsA.cell(row=row, column=1).value is None: 
        continue

    dealName = wsA.cell(row=row, column=3).value.split('-')[0]
    if dealName in seen: 
        continue

    seen.add(dealName)

    for mapping in mappings:
        colA = Utils.toIndex(mapping[0]) if len(mapping[0]) == 1 else Utils.toIndexes(mapping[0])
        colB = Utils.toIndex(mapping[1]) if len(mapping[1]) == 1 else Utils.toIndexes(mapping[1])
        match mapping[2]:
            case Copy.NONE:
                wsB.cell(row=row + OFFSET, column=colB).value = None
            case Copy.VALUE:
                wsB.cell(row=row + OFFSET, column=colB).value = wsA.cell(row=row, column=colA).value
            case Copy.YES:
                wsB.cell(row=row + OFFSET, column=colB).value = 'Y'
            case Copy.NO:
                wsB.cell(row=row + OFFSET, column=colB).value = 'N'
            case Copy.DSAR:
                wsB.cell(row=row + OFFSET, column=colB).value = 'DSAR'
            case _:
                continue

    for street in streets:
        # Column in Input Spreadsheet
        colA = Utils.toIndex(street[0])
        # Full Address from Input Spreadsheet, if invalid, skip
        fullAddress = wsA.cell(row=row, column=colA).value
        if not fullAddress or fullAddress is None or fullAddress == "(No Value)" or len(str(fullAddress).strip()) < 11:
            continue

        # Number of Address Lines to get from streets mapping
        linesAddress = len(street[1])
        # Split full address into parts for each line
        split = fullAddress.rsplit(',', linesAddress - 1)
        # print(f'split: {len(split)}, row: {row}, fA: {fullAddress}, fAL: {len(fullAddress)}') # debugging

        # If not enough commas, divide less
        if not len(split) == linesAddress:
            linesAddress = len(split)

        # Get lines of address from split
        for i in range (0, linesAddress):
            column = Utils.toIndex(street[1][i])
            wsB.cell(row=row + OFFSET, column=column).value = split[i].upper() if i == 2 else split[i].strip()

# Save and Close
wb2.save(OP_FILENAME)
wb1.close()
wb2.close()