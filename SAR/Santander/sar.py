# Run: Python3 sar.py

import openpyxl
from openpyxl import load_workbook

wb1 = load_workbook('SAR_Santander_HS_2504.xlsx')
wsA = wb1['SAR_Santander']
# wsB = wb['output']

wb2 = load_workbook('Santander_Template.xlsx')
wsB = wb2['Submission Template']

# clean up the output sheet
for row in wsB.iter_rows(min_row=2, max_col=36, max_row=350):
    for cell in row:
        cell.value = None

# for i in range(1, 11):
#     wsB[f"A{i}"] = wsA[f"A{i}"].value

# [A1, B2, C3, D4, E5, F6, G7, H8, I9, J10, K11]
# [L12, M13, N14, O15, P16, Q17, R18, S19, T20, U21, V22, W23, X24]

# range -> [a, b)
for row in range(2, 325):  # 2 to 3
    # NAME
    for column in range(1, 3):  # A to B
        wsB.cell(row=row, column=column).value = wsA.cell(row=row, column=column + 3).value
    # PNAMES
    for column in range(3, 6):  # C to E
        wsB.cell(row=row, column=column).value = None
    # DOB
    for column in range(6, 7):  # F
        wsB.cell(row=row, column=column + 3).value = wsA.cell(row=row, column=column).value
    for column in range(7, 9):
        # i = 0 for first column, i = 1 for second column
        i = column - 7
        fulladdress = wsA.cell(row=row, column=column).value
        if fulladdress is not None:
            # split on the last comma
            split = fulladdress.rsplit(',', 1)
            if len(split) > 1:
                # Address
                wsB.cell(row=row, column=column + 3 + i).value = split[0]
                # Postcode
                wsB.cell(row=row, column=column + 4 + i).value = split[1]
        elif column == 8 and fulladdress is None:
            wsB.cell(row=row, column=column + 4).value = wsB.cell(row=row, column=column + 2).value
            wsB.cell(row=row, column=column + 5).value = wsB.cell(row=row, column=column + 3).value
    for column in range(9, 12):
        cell = wsA.cell(row=row, column=column).value
        if cell is not None:
            wsB.cell(row=row, column = column + 13).value = cell
        # otherwise, leave it blank

wb2.save('template.xlsx')