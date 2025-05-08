# Run: Python3 sar.py

import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

wb1 = load_workbook('Stellantis-Toyota-2025-04-28.xlsx')
wsA = wb1['SAR_Default']

wb2 = load_workbook('Stellantis.xlsx')
wsB = wb2['Sheet1']

# clean up the output sheet
for row in wsB.iter_rows(min_row=2, max_col=25, max_row=390):
    for cell in row:
        cell.value = None

# for i in range(1, 11):
#     wsB[f"A{i}"] = wsA[f"A{i}"].value

# [A1, B2, C3, D4, E5, F6, G7, H8, I9, J10, K11]
# [L12, M13, N14, O15, P16, Q17, R18, S19, T20]
# [U21, V22, W23, X24, Y25, Z26]

# range -> [a, b)
for row in range(2, 390):  # 2 to 390
    if wsA.cell(row=row, column=1).value is None : continue
    for column in range(1, 2):  # A to B
        wsB.cell(row=row, column=column).value = wsA.cell(row=row, column=column).value
    # NAME
    for column in range(4, 6):  # A to B
        wsB.cell(row=row, column=column - 2).value = wsA.cell(row=row, column=column).value
    # DOB
    for column in range(6, 7):  # C to E
        wsB.cell(row=row, column=column - 1).value = wsA.cell(row=row, column=column).value
    # ADDRESS
    for column in range(7, 9):
        # i = 0 for first column, i = 1 for second column
        i = column - 7
        fulladdress = wsA.cell(row=row, column=column).value
        # split on the last comma
        split = fulladdress.rsplit(',', 1)
        if fulladdress is not None and len(split) > 1:
            # Address
            wsB.cell(row=row, column=column - 1 + i).value = split[0]
            # Postcode
            wsB.cell(row=row, column=column + i).value = split[1]
        # elif column == 8 and fulladdress is None:
        #     wsB.cell(row=row, column=column + 4).value = wsB.cell(row=row, column=column + 2).value
        #     wsB.cell(row=row, column=column + 5).value = wsB.cell(row=row, column=column + 3).value
    for column in range(9, 10):
        cell = wsA.cell(row=row, column=column).value
        if cell is not None:
            wsB.cell(row=row, column = column + 1).value = cell
        # otherwise, leave it blank

wb2.save('Stellantis.xlsx')

# def copy_row(wsA, wsB, row, col_A, col_B, isNone):
#     return 0