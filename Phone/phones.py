# Run: Python3 phones.py
import openpyxl                     # type: ignore
from openpyxl import load_workbook  # type: ignore

unseen = set()

# 1. open text file with phone numbers
with open("phones.txt", 'r') as f:
    for line in f:
        unseen.add(line.strip())

seen = set()

output = "RecordID,Phone,Provider,PartnerID,Signed(Y/N)\n"

# 2. open excel file with all contacts
FILENAME = 'all-contacts.xlsx'
REC_COL = 1
PHONE_COL = 5

wb = load_workbook(FILENAME)
ws = wb.active

# 3. for each phone number, attempt to find in contact registry

# for phone in unseen:
    # print(phone)

row = 2
while True:
    record = ws.cell(row=row, column=REC_COL).value
    contact_phone = ws.cell(row=row, column=PHONE_COL).value

    if record is None or record == "":
        break

    if contact_phone is None or contact_phone == "":
        row += 1
        continue

    for phone in list(unseen):
        if phone in str(contact_phone):
            seen.add(phone)
            unseen.remove(phone)
            print(f"Found {phone} in record {record} with phone {contact_phone}")

            # Provider
            op1 = ws.cell(row=row, column=PHONE_COL + 1).value

            # PartnerID
            op2 = ws.cell(row=row, column=PHONE_COL + 2).value

            # Live status
            op3 = ws.cell(row=row, column=PHONE_COL + 3).value
            op4 = ws.cell(row=row, column=PHONE_COL + 4).value

            op5 = "Y" if op3 == "LOA file" or op4 == "Yes" else "N"

            output += f"{record},{contact_phone},{op1},{op2},{op5}\n"

    row += 1

print(f"rows: {row}")

# # 4. write output to csv
with open("phones-out.csv", 'w') as f:
    f.write(output)

with open("phones-not-found.txt", 'w') as f:
    for phone in unseen:
        f.write(f"{phone}\n")

# 5. for found, output all things, including name, phone, etc.
wb.close