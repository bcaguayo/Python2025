# Run: Python3 presubglobal.py

from openpyxl import load_workbook  #type:ignore
from enum import Enum
from utils import Utils

IP_FILENAME = 'src/testA.xlsx'
IP_SHEETNAME = 'Sheet1'

OP_FILENAME = 'templates/testB.xlsx'
OP_SHEETNAME = 'Sheet1'

Copy = Enum('Copy', [('NONE', 0), ('VALUE', 1)])


# Load These
config = {
  'minRow' : 1,
  'maxRow' : 11
}

mappings = {
  "a": 1,
  "z": 0,
  "y": 0,
  "k": 1,
  "b": 1,
  "c": 1
}

# Init Src Book
wbA = load_workbook(IP_FILENAME)
wsA = wbA[IP_SHEETNAME]

# Init Snk Book
wbB = load_workbook(OP_FILENAME)
wsB = wbB[OP_SHEETNAME]

# Get Config
minRows = config['minRow']
maxRows = config['maxRow']

for row in range(minRows, maxRows):
  # mappings replace columns
  for idx, mapping in enumerate(mappings):
    source_col = idx
    target_col = Utils.toIndex(mapping)
    target_enum = mappings[mapping]


# Loop for different files
# Go through file
# Copy values based on mappings
# Get utils

