# Run: Python3 parts.py

# Import math library
import math
import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

# init values
NUM_COLUMNS = 9
NUM_ROWS = 192 # (include header)
NUM_PARTITIONS = 8

# Input -> filename and sheet name
FILENAME = 'lost.xlsx'
SHEETNAME = 'Sheet1'

# Calculate number of rows per partition
parts = math.ceil(NUM_ROWS / NUM_PARTITIONS)

# Open Workbook
wb = load_workbook(FILENAME)
ws = wb[SHEETNAME]
ws_parts = []

# Make NUM_PARTITIONS sheets
for i in range(1, NUM_PARTITIONS + 1):
    ws_new = wb.create_sheet(f'Part {i}')
    ws_parts.append(ws_new)

# Stash header
header = []
num_columns = 1
cell = ws.cell(row=1, column=num_columns).value
while cell != None:
    header.append(cell)
    num_columns += 1
    cell = ws.cell(row=1, column=num_columns).value

# to make up for the additional value in loop
num_columns -= 1

# Loopity Loop
i = 0       # index on local partition
i_part = 0  # index on array of sheets
ws_active = ws_parts[0]

# Split data into n partitions
for row in range (2, NUM_ROWS + 1):
    if i > parts:
        i = 0
        i_part += 1
        ws_active = ws_parts[i_part]
        for column in range (1, num_columns + 1):
            ws_active.cell(row=i+1, column=column).value = header[column - 1]
    else:
        for column in range (1, num_columns + 1):
            ws_active.cell(row=i+1, column=column).value = ws.cell(row=row, column=column).value
    i += 1
    
# Womp womp
wb.save(FILENAME)
wb.close