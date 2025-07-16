from openpyxl import load_workbook      # type: ignore

FILENAME = 'dates2.xlsx'

wb = load_workbook(FILENAME)
ws = wb.active

# Counting Rows and Predef Values
# rows = 1
COL_START_DATE = 4
COL_END_DATE = 5
COL_IN_DATE = 6
ROW_LIMIT = 2

while True :
    record = ws.cell(row=ROW_LIMIT,column=1).value
    if record is None or record == "" : 
        break
    # ws.cell(row=ROW_LIMIT, column=COL_IN_DATE).value = ""
    ROW_LIMIT += 1

for row in range(2, ROW_LIMIT) :

    start_date = ws.cell(row=row,column=COL_START_DATE).value
    end_date = ws.cell(row=row,column=COL_END_DATE).value

    # Veredict -> In Date Column
    # In Date: yes, no - before 01/04/2007, no - after 28/01/2021, blank
    # Y, N2007 or N07, N2021 or N21, B    

    if start_date is None or start_date == "" and end_date is None or end_date == "" :
        # print(f'blank on row: {rows}')
        ws.cell(row=row,column=COL_IN_DATE).value = "B"
    
    else :
        # Start Date before 01/04/2007
        sd_pre_2007 = False
        # Start Date after 28/01/2021
        sd_post_2021 = False
        # End Date before 01/04/2007
        ed_pre_2007 = False
        # End Date before 28/01/2021
        ed_pre_2021 = False
        # End Date before 04/12/2025
        ed_pre_2025 = False
        # Blank Start Date
        blankSD = start_date is None or start_date == "" 
        # Blank End Date
        blankED = end_date is None or end_date == ""

        if not blankSD :
            day = start_date.day
            month = start_date.month
            year = start_date.year
            # Before 01/04/2007
            if year < 2004 or year == 2004 and month < 4 :
                sd_pre_2007 = True
            # After 28/01/2021
            elif year > 2021 or year == 2021 and month > 1 or year == 2021 and month == 1 and day > 28 :
                sd_post_2021 = True

        if not blankED :
            day = end_date.day
            month = end_date.month
            year = end_date.year
            # Before 01/04/2007
            if year < 2004 or year == 2004 and month < 4 :
                ed_pre_2007 = True
            # Before 28/01/2021            
            elif year < 2021 or year == 2021 and month < 1 or year == 2021 and month == 1 and day < 28 :
                ed_pre_2021 = True
            # Before 04/12/2025  
            elif year < 2025 or year == 2025 and month < 12 or year == 2025 and month == 12 and day < 4 :
                ed_pre_2025 = True
        
        if blankSD or blankED :
            ws.cell(row=row, column=COL_IN_DATE).value = "B"
        
        elif sd_pre_2007 and (ed_pre_2007 or ed_pre_2025) :
            ws.cell(row=row, column=COL_IN_DATE).value = "Pre 2007"
        
        elif sd_post_2021 and (not ed_pre_2025 or blankED) :
            ws.cell(row=row, column=COL_IN_DATE).value = "After 2021"
            
        elif sd_pre_2007 and ed_pre_2021 :
            ws.cell(row=row, column=COL_IN_DATE).value = "Y07"

        elif sd_post_2021 and ed_pre_2025 :
            ws.cell(row=row, column=COL_IN_DATE).value = "Y21"

        else :
            ws.cell(row=row, column=COL_IN_DATE).value = "Y"

wb.save(FILENAME)
wb.close()