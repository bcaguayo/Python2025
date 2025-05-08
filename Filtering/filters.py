# Run: Python3 sar.py

import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

FILENAME = 'SAR_Novuna.xlsx'
SHEETNAME = 'SAR_Novuna'

wb1 = load_workbook(FILENAME)
wsA = wb1[SHEETNAME]

NUM_ROWS = 180
LOA_COL = 12
LOA_FIL_COL = LOA_COL + 1
SPLIT = True
SORTED = True

if not SPLIT:
    for row in range(2, NUM_ROWS):  # 2 to 105
        loa = wsA.cell(row=row, column=LOA_COL).value
        if loa != None : 
            wsA.cell(row=row, column=LOA_FIL_COL).value = loa.split(')', 1)[-1].strip()

# Count the number of non-empty cells in column 14 if not duplicated
seen = set()
if SORTED:
    for row in range(2, 30):  # 2 to 105
        loa = wsA.cell(row=row, column=LOA_FIL_COL).value
        if loa != None and loa not in seen: 
            seen.add(loa)
        elif loa in seen:
            wsA.cell(row=row, column=LOA_FIL_COL).value = None

print("Number of unique non-empty cells in column 14:", len(seen))

wb1.save(FILENAME)
wb1.close()