from openpyxl import load_workbook

FILENAME = 'autobank.xlsx'

wb = load_workbook(FILENAME)
wsA = wb['db']
wsB = wb['res']

AFAN_INDEX = 5
BFAN_INDEX = 2
NUM_AFAN = 305
COL_RID = 1
COL_NAME = 2
COL_STAGE = 3

db = list()
db.append(0)
db.append(1)
for i in range(2, NUM_AFAN + 1) :
    afan = str(wsA.cell(row=i, column=5).value).strip(' ')
    db.append(afan)

# for fan in db:
#     print(f'item {db.index(fan)}: {fan}, type: {type(fan)}')

# Locate FAN from wsB in wsA and copy over cols A, B, C to wsB
for i in range(2, 293) :
    bfan = str(wsB.cell(row=i, column=BFAN_INDEX).value).strip(' ')
    j = db.index(bfan) if bfan in db else -1
    if j != -1 :
        # afan = wsA.cell(row=j, column=AFAN_INDEX).value
        # print(f'{afan}:{bfan}')
        rid = wsA.cell(row=j, column=COL_RID).value
        name = wsA.cell(row=j, column=COL_NAME).value
        stage = wsA.cell(row=j, column=COL_STAGE).value
        wsB.cell(row=i, column=9 + COL_RID).value = rid
        wsB.cell(row=i, column=9 + COL_NAME).value = name
        wsB.cell(row=i, column=9 + COL_STAGE).value = stage
    else :
        wsB.cell(row=i, column=10).value = 'not found'
        # print(bfan)
        # print(f'item at {i}: {bfan}, type: {type(bfan)}, not found')

wb.save(FILENAME)
wb.close()