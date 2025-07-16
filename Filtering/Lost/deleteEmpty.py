# Global build python3 -m pip install openpyxl
# Run: Python3 lost.py

# Import math library
import math
import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

# Input -> filename and sheet name
FILENAME = 'lost.xlsx'
SHEETNAME = 'manual'

# Open Workbook
wb = load_workbook(FILENAME)
ws = wb[SHEETNAME]

# init values
ROW_LIMIT = 53156

for row in range(ROW_LIMIT, 1, -1):
    print(f'row num: {row}')
    if ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value == "":
        print(f'deleted: {row}')
        ws.delete_rows(row, 1)

wb.save(FILENAME)
wb.close()