# Global build python3 -m pip install openpyxl
# Run: Python3 lost.py

# Import math library
import math
import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

# Input -> filename and sheet name
FILENAME = 'lost - Dupes.xlsx'
SHEETNAME = 'clear'

# Open Workbook
wb = load_workbook(FILENAME)
ws = wb[SHEETNAME]

# init values
ROW_LIMIT = 2
CONTACT_COL = 7
NUM_COLUMNS = 8

seen = set()
count = 0

while True :
    # EOF
    if ws.cell(row=ROW_LIMIT,column=1).value is None or ws.cell(row=ROW_LIMIT,column=1).value == "" : 
        break

    # SEEN
    record = ws.cell(row=ROW_LIMIT,column=CONTACT_COL).value
    if record in seen :
        for col in range(1, NUM_COLUMNS + 1) :
            ws.cell(row=ROW_LIMIT,column=col).value = ""
        count += 1
    else :
        seen.add(record)

    # RAISE EOF LIMIT
    ROW_LIMIT += 1

print(f'deleted: {count}')

wb.save(FILENAME)
wb.close()