# Global build python3 -m pip install openpyxl
# Run: Python3 lost.py

# Import math library
import math
import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

# Input -> filename and sheet name
FILENAME = 'lost.xlsx'
SHEETNAME = 'manual'

# Open Workbook
wb = load_workbook(FILENAME)
ws = wb[SHEETNAME]

# init values
NUM_COLUMNS = 8
ROW_LIMIT = 2
CONTACT_COL = 7
LOST_COL = 2

notlost = set()

while True :
    # EOF
    if ws.cell(row=ROW_LIMIT,column=1).value is None or ws.cell(row=ROW_LIMIT,column=1).value == "" : 
        break

    # SEEN
    lost = ws.cell(row=ROW_LIMIT,column=LOST_COL).value
    if "Lost" not in str(lost) :
        # Add Contact not in Lost
        notlost.add(ws.cell(row=ROW_LIMIT,column=CONTACT_COL).value)

    # RAISE EOF LIMIT
    ROW_LIMIT += 1

# Make not Lost blank
for row in range(2, ROW_LIMIT):
    if ws.cell(row=row,column=CONTACT_COL).value in notlost:
        # lost = ws.cell(row=row,column=LOST_COL).value
        # name = str(ws.cell(row=row,column=3).value) + ' ' +  str(ws.cell(row=row,column=4).value)
        # print(f'row: {row}, stage: {lost}, contact: {name}')
        # empty the row
        for col in range(1, NUM_COLUMNS + 1) :
            ws.cell(row=row,column=col).value = ""

wb.save(FILENAME)
wb.close()