# Run: Python3 sar.py

import datetime
import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore

FILENAME = 'plare.xlsx'
SHEETNAME = 'PLARE Enquiry'

wb1 = load_workbook(FILENAME)
wsA = wb1[SHEETNAME]

NUM_ROWS = 180
DATE_A_COL = 2
DATE_B_COL = 3
TIME_L_COL = 4

discrepancies = set()

for row in range(2, NUM_ROWS):
    # 1. Get Name
    record = wsA.cell(row=row, column=1).value
    if record == None : break

    # 2. Get Dates of Creation
    dateA = wsA.cell(row=row, column=DATE_A_COL).value
    dateB = wsA.cell(row=row, column=DATE_B_COL).value
    # timedelta between dateA and dateB
    delta =  dateB - dateA
    # fifteen = fifteen minutes
    fifteen = datetime.timedelta(minutes=15)

    # 3. Get Time on Lead
    time = wsA.cell(row=row, column=TIME_L_COL).value
    # timeL = wsA.cell(row=row, column=TIME_L_COL).value if time != None else datetime.timedelta(minutes=0)
    hours, mins, secs = map(int, time.split(':'))
    # print(f'hours: {hours} - over 0 {hours > 0}, mins {mins} - over 15 {mins>=15}')
    over15 = hours > 0 or mins >= 15

    # B(Deal) should always be bigger than A(Client)
    if (delta <= fifteen and not over15 and not record in discrepancies) : # or delta > -fifteen) :        
        discrepancies.add(record)
        print(f'added: {record}, with dif: {delta}')
# profit

# wb1.save(FILENAME)
wb1.close()